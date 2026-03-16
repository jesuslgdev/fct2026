from io import BytesIO

from openpyxl import load_workbook

from modules.suppliers.application.download_supplier_template_use_case import (
    DownloadSupplierTemplateUseCase,
)


def test_template_returns_bytes():
    result = DownloadSupplierTemplateUseCase().execute()
    assert isinstance(result, bytes)
    assert len(result) > 0


def test_template_sheet_title():
    result = DownloadSupplierTemplateUseCase().execute()
    wb = load_workbook(BytesIO(result))
    assert wb.active.title == "Proveedores"


def test_template_headers():
    result = DownloadSupplierTemplateUseCase().execute()
    wb = load_workbook(BytesIO(result))
    ws = wb.active
    headers = [
        ws.cell(1, col).value
        for col in range(1, len(DownloadSupplierTemplateUseCase.HEADERS) + 1)
    ]
    assert headers == list(DownloadSupplierTemplateUseCase.HEADERS)


def test_template_example_row():
    result = DownloadSupplierTemplateUseCase().execute()
    wb = load_workbook(BytesIO(result))
    ws = wb.active
    example = [
        ws.cell(2, col).value
        for col in range(1, len(DownloadSupplierTemplateUseCase.EXAMPLE) + 1)
    ]
    assert example == list(DownloadSupplierTemplateUseCase.EXAMPLE)
