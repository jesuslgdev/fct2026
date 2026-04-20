from io import BytesIO
from unittest.mock import AsyncMock, MagicMock

import pytest
from openpyxl import load_workbook

from modules.suppliers.application.download_supplier_product_template_use_case import (
    DownloadSupplierProductTemplateUseCase,
)
from modules.suppliers.domain.exceptions import (
    SupplierException,
    SupplierExceptionInfo,
)


def _worksheet_rows(content: bytes) -> list[tuple]:
    wb = load_workbook(BytesIO(content))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return rows


@pytest.mark.asyncio
async def test_template_returns_example_row_without_selected_products():
    reader = MagicMock()
    reader.get_by_id = AsyncMock()
    use_case = DownloadSupplierProductTemplateUseCase(reader)

    result = await use_case.execute()
    rows = _worksheet_rows(result)

    assert rows[0] == DownloadSupplierProductTemplateUseCase.HEADERS
    assert rows[1] == DownloadSupplierProductTemplateUseCase.EXAMPLE
    reader.get_by_id.assert_not_called()


@pytest.mark.asyncio
async def test_template_prefills_selected_product_codes():
    product_1 = MagicMock(product_code="PROD-001", is_active=True)
    product_1.name = "Producto 1"
    product_2 = MagicMock(product_code="PROD-002", is_active=True)
    product_2.name = "Producto 2"
    products = {1: product_1, 2: product_2}

    reader = MagicMock()
    reader.get_by_id = AsyncMock(
        side_effect=lambda product_id: products.get(product_id)
    )
    use_case = DownloadSupplierProductTemplateUseCase(reader)

    result = await use_case.execute(product_ids=[2, 1, 2])
    rows = _worksheet_rows(result)

    assert rows[0] == DownloadSupplierProductTemplateUseCase.HEADERS
    assert rows[1] == ("PROD-002", None, "Producto 2")
    assert rows[2] == ("PROD-001", None, "Producto 1")


@pytest.mark.asyncio
async def test_template_raises_when_product_not_found():
    reader = MagicMock()
    reader.get_by_id = AsyncMock(return_value=None)
    use_case = DownloadSupplierProductTemplateUseCase(reader)

    with pytest.raises(SupplierException) as exc:
        await use_case.execute(product_ids=[999])

    assert exc.value.info == SupplierExceptionInfo.PRODUCT_NOT_FOUND


@pytest.mark.asyncio
async def test_template_raises_when_product_is_inactive():
    inactive_product = MagicMock(product_code="PROD-099", is_active=False)
    inactive_product.name = "Producto 99"
    reader = MagicMock()
    reader.get_by_id = AsyncMock(return_value=inactive_product)
    use_case = DownloadSupplierProductTemplateUseCase(reader)

    with pytest.raises(SupplierException) as exc:
        await use_case.execute(product_ids=[99])

    assert exc.value.info == SupplierExceptionInfo.PRODUCT_NOT_ACTIVE
