import re
from io import BytesIO

from openpyxl import load_workbook

from modules.suppliers.domain.entities.import_result import (
    ImportResult,
    ImportRowError,
)
from modules.suppliers.domain.entities.supplier import Supplier
from modules.suppliers.domain.interfaces.repositories.i_supplier_repository import (
    ISupplierRepository,
)
from modules.suppliers.domain.interfaces.use_cases.i_import_suppliers_use_case import (
    IImportSuppliersUseCase,
)
from shared.constants import (
    EMAIL_PATTERN,
    PHONE_PATTERN,
    POSTAL_CODE_PATTERN,
    TAX_ID_PATTERN,
)

COLUMN_MAP = {
    "Nombre": "name",
    "CIF": "tax_id",
    "Dirección": "address",
    "Ciudad": "city",
    "Provincia": "province",
    "Código Postal": "postal_code",
    "Teléfono": "phone",
    "Email": "email",
}
EXPECTED_HEADERS = list(COLUMN_MAP.keys())
CIF_REGEX = re.compile(TAX_ID_PATTERN)
EMAIL_REGEX = re.compile(EMAIL_PATTERN)
POSTAL_CODE_REGEX = re.compile(POSTAL_CODE_PATTERN)
PHONE_REGEX = re.compile(PHONE_PATTERN)

MAX_LENGTHS = {
    "name": 150,
    "address": 255,
    "city": 100,
    "province": 100,
    "email": 150,
}


class ImportSuppliersUseCase(IImportSuppliersUseCase):
    def __init__(self, repo: ISupplierRepository) -> None:
        self._repo = repo

    async def execute(self, file_content: bytes) -> ImportResult:
        try:
            wb = load_workbook(BytesIO(file_content), read_only=True, data_only=True)
        except Exception:
            return ImportResult(
                total=0,
                created=0,
                errors=[ImportRowError(row=1, reason="Invalid or corrupted file")],
            )

        try:
            ws = wb.active
            row_iter = ws.iter_rows(min_row=1, values_only=True)

            header_row = next(row_iter, None)
            if header_row is None:
                return ImportResult(
                    total=0,
                    created=0,
                    errors=[
                        ImportRowError(
                            row=1, reason="File is empty or has no data rows"
                        )
                    ],
                )

            headers = [str(h).strip() if h else "" for h in header_row]
            header_prefix = headers[: len(EXPECTED_HEADERS)]
            extra_headers = headers[len(EXPECTED_HEADERS) :]
            if header_prefix != EXPECTED_HEADERS or any(h for h in extra_headers):
                return ImportResult(
                    total=0,
                    created=0,
                    errors=[
                        ImportRowError(
                            row=1, reason="Invalid headers. Use the provided template"
                        )
                    ],
                )

            errors: list[ImportRowError] = []
            parsed: list[tuple[int, Supplier]] = []
            seen_tax_ids: set[str] = set()
            total = 0

            for i, row in enumerate(row_iter, start=2):
                total += 1
                if not any(row):
                    total -= 1
                    continue
                row_errors = self._validate_row(row, i, seen_tax_ids)
                if row_errors:
                    errors.extend(row_errors)
                else:
                    fields = {
                        db_field: str(row[col_idx]).strip()
                        for col_idx, db_field in enumerate(COLUMN_MAP.values())
                    }
                    fields["tax_id"] = fields["tax_id"].upper()
                    parsed.append((i, Supplier(**fields)))
                    seen_tax_ids.add(fields["tax_id"])

            if parsed and not errors:
                existing = await self._repo.get_existing_tax_ids(
                    [s.tax_id for _, s in parsed]
                )
                for row_num, s in parsed:
                    if s.tax_id in existing:
                        errors.append(
                            ImportRowError(
                                row=row_num,
                                reason=f"CIF {s.tax_id} already exists in database",
                            )
                        )

            if errors:
                return ImportResult(total=total, created=0, errors=errors)

            created = await self._repo.bulk_create([s for _, s in parsed])
            return ImportResult(total=total, created=created, errors=[])
        finally:
            wb.close()

    def _validate_row(
        self, row: tuple, row_num: int, seen_tax_ids: set[str]
    ) -> list[ImportRowError]:
        errors: list[ImportRowError] = []
        field_names = list(COLUMN_MAP.keys())
        field_keys = list(COLUMN_MAP.values())

        for col_idx, header in enumerate(field_names):
            value = row[col_idx] if col_idx < len(row) else None
            str_value = str(value).strip() if value is not None else ""
            if str_value == "":
                errors.append(
                    ImportRowError(row=row_num, reason=f"Field '{header}' is required")
                )
                continue

            field_key = field_keys[col_idx]
            if field_key in MAX_LENGTHS and len(str_value) > MAX_LENGTHS[field_key]:
                errors.append(
                    ImportRowError(
                        row=row_num,
                        reason=f"Field '{header}' exceeds maximum length of {MAX_LENGTHS[field_key]}",
                    )
                )

        if errors:
            return errors

        tax_id = str(row[1]).strip().upper()
        postal_code = str(row[5]).strip()
        phone = str(row[6]).strip()
        email = str(row[7]).strip()

        if not CIF_REGEX.match(tax_id):
            errors.append(ImportRowError(row=row_num, reason="Invalid CIF format"))

        if not POSTAL_CODE_REGEX.match(postal_code):
            errors.append(
                ImportRowError(row=row_num, reason="Invalid postal code format")
            )

        if not PHONE_REGEX.match(phone):
            errors.append(ImportRowError(row=row_num, reason="Invalid phone format"))

        if not EMAIL_REGEX.match(email):
            errors.append(ImportRowError(row=row_num, reason="Invalid email format"))

        if tax_id in seen_tax_ids:
            errors.append(ImportRowError(row=row_num, reason="Duplicate CIF in file"))

        return errors
