from decimal import Decimal, InvalidOperation
from io import BytesIO

from openpyxl import load_workbook

from modules.suppliers.domain.dtos.import_result import (
    ImportResult,
    ImportRowError,
)
from modules.suppliers.domain.entities.supplier_product import SupplierProduct
from modules.suppliers.domain.exceptions import SupplierException, SupplierExceptionInfo
from modules.suppliers.domain.interfaces.repositories.i_supplier_repository import (
    ISupplierRepository,
)
from modules.suppliers.domain.interfaces.use_cases.i_import_supplier_products_use_case import (
    IImportSupplierProductsUseCase,
)
from shared.domain.interfaces.i_product_reader import IProductReader

EXPECTED_HEADERS = ["Código Producto", "Precio Proveedor"]


class ImportSupplierProductsUseCase(IImportSupplierProductsUseCase):
    def __init__(
        self, repo: ISupplierRepository, product_reader: IProductReader
    ) -> None:
        self._repo = repo
        self._product_reader = product_reader

    async def execute(self, supplier_id: int, file_content: bytes) -> ImportResult:
        supplier = await self._repo.get_by_id(supplier_id)
        if not supplier:
            raise SupplierException(SupplierExceptionInfo.SUPPLIER_NOT_FOUND)
        if not supplier.is_active:
            raise SupplierException(SupplierExceptionInfo.SUPPLIER_NOT_ACTIVE)

        try:
            wb = load_workbook(BytesIO(file_content), read_only=True, data_only=True)
        except Exception:
            return ImportResult(
                total=0,
                created=0,
                errors=[ImportRowError(row=1, reason="Invalid or corrupted file")],
            )

        try:
            ws = wb.active
            row_iter = ws.iter_rows(min_row=1, values_only=True)

            header_row = next(row_iter, None)
            if header_row is None:
                return ImportResult(
                    total=0,
                    created=0,
                    errors=[
                        ImportRowError(
                            row=1, reason="File is empty or has no data rows"
                        )
                    ],
                )

            headers = [str(h).strip() if h else "" for h in header_row]
            header_prefix = headers[: len(EXPECTED_HEADERS)]
            extra_headers = headers[len(EXPECTED_HEADERS) :]
            if header_prefix != EXPECTED_HEADERS or any(h for h in extra_headers):
                return ImportResult(
                    total=0,
                    created=0,
                    errors=[
                        ImportRowError(
                            row=1, reason="Invalid headers. Use the provided template"
                        )
                    ],
                )

            errors: list[ImportRowError] = []
            parsed: list[tuple[int, str, Decimal]] = []
            seen_product_codes: set[str] = set()
            total = 0

            for i, row in enumerate(row_iter, start=2):
                total += 1
                if not any(row):
                    total -= 1
                    continue
                row_errors = self._validate_row(row, i, seen_product_codes)
                if row_errors:
                    errors.extend(row_errors)
                else:
                    product_code = str(row[0]).strip()
                    price = Decimal(str(row[1]))
                    parsed.append((i, product_code, price))
                    seen_product_codes.add(product_code)

            to_create: list[SupplierProduct] = []

            if parsed and not errors:
                for row_num, product_code, price in parsed:
                    product = await self._product_reader.get_by_code(product_code)
                    if not product:
                        errors.append(
                            ImportRowError(
                                row=row_num,
                                reason=f"Product with code {product_code} not found",
                            )
                        )
                        continue
                    if not product.is_active:
                        errors.append(
                            ImportRowError(
                                row=row_num,
                                reason=f"Product {product_code} is not active",
                            )
                        )
                        continue

                    existing = await self._repo.get_association(
                        supplier_id, product.product_id
                    )
                    if existing:
                        errors.append(
                            ImportRowError(
                                row=row_num,
                                reason=f"Association already exists for product {product_code}",
                            )
                        )
                        continue

                    to_create.append(
                        SupplierProduct(
                            supplier_id=supplier_id,
                            product_id=product.product_id,
                            supplier_price=price,
                        )
                    )

            if errors:
                return ImportResult(total=total, created=0, errors=errors)

            created = await self._repo.bulk_create_products(to_create)
            return ImportResult(total=total, created=created, errors=[])
        finally:
            wb.close()

    def _validate_row(
        self, row: tuple, row_num: int, seen_product_codes: set[str]
    ) -> list[ImportRowError]:
        errors: list[ImportRowError] = []

        product_code_val = row[0] if len(row) > 0 else None
        price_val = row[1] if len(row) > 1 else None

        product_code = str(product_code_val).strip() if product_code_val else ""
        if not product_code:
            errors.append(
                ImportRowError(row=row_num, reason="Product code is required")
            )

        if price_val is None:
            errors.append(ImportRowError(row=row_num, reason="Price is required"))

        if errors:
            return errors

        try:
            price = Decimal(str(price_val))
            if price <= 0:
                errors.append(
                    ImportRowError(
                        row=row_num, reason="Price must be greater than zero"
                    )
                )
            elif price.as_tuple().exponent < -2:
                errors.append(
                    ImportRowError(
                        row=row_num,
                        reason="Price cannot have more than 2 decimal places",
                    )
                )
        except (InvalidOperation, ValueError):
            errors.append(ImportRowError(row=row_num, reason="Invalid price format"))

        if product_code in seen_product_codes:
            errors.append(
                ImportRowError(
                    row=row_num,
                    reason=f"Duplicate product code in file: {product_code}",
                )
            )

        return errors
